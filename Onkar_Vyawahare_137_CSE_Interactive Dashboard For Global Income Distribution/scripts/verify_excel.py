import openpyxl
import os
import random

def verify_excel_values(filename):
    """
    Verifies that floating-point numbers in the given Excel file are rounded to 2 decimal places.
    """
    try:
        if not os.path.exists(filename):
            print(f"Error: File '{filename}' not found.")
            return

        print(f"Loading workbook for verification: {filename}")
        wb = openpyxl.load_workbook(filename, data_only=True) # data_only=True to get values, not formulas

        count_checked = 0
        count_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Checking sheet: {sheet_name}")
            
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, float):
                        count_checked += 1
                        # Check if more than 2 decimal places
                        # Note: due to float precision, we check string representation or small epsilon
                        # A robust check: round(val, 2) == val
                        if round(cell.value, 2) != cell.value:
                            # Allow for tiny floating point errors, e.g. 1.23000000001
                            if abs(round(cell.value, 2) - cell.value) > 1e-9:
                                print(f"  Wait! Found unrounded value at {cell.coordinate}: {cell.value}")
                                count_errors += 1
                                if count_errors > 5:
                                    print("  Too many errors, stopping check for this sheet.")
                                    break
        
        if count_checked == 0:
            print("Warning: No floating point values found to check.")
        elif count_errors == 0:
            print(f"Verification Success! Checked {count_checked} float values. All appear correctly rounded.")
        else:
            print(f"Verification Failed! Found {count_errors} unrounded values out of {count_checked} checked.")

    except Exception as e:
        print(f"An verification error occurred: {e}")

if __name__ == "__main__":
    input_file = "Mastersheet_Infosys_Rounded.xlsx"
    base_path = r"c:\Users\onkar\Desktop\infosys springboard virtual internship (1)"
    full_input_path = os.path.join(base_path, input_file)
    
    verify_excel_values(full_input_path)
