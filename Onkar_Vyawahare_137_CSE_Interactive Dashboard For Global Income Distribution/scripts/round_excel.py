import openpyxl
import os

def round_excel_values(filename):
    """
    Rounds all floating-point numbers in the given Excel file to 2 decimal places.
    Preserves formatting.
    Saves the result to a new file with '_Rounded' suffix.
    """
    try:
        if not os.path.exists(filename):
            print(f"Error: File '{filename}' not found.")
            return

        print(f"Loading workbook: {filename}")
        wb = openpyxl.load_workbook(filename)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Processing sheet: {sheet_name}")
            
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (float, int)) and not isinstance(cell.value, bool):
                        # check if it is a float and round it
                        if isinstance(cell.value, float):
                             cell.value = round(cell.value, 2)
                        # Identify if it was an int but maybe represented as float in some cases? 
                        # The user specifically mentioned "55017.3049786608 keep only 55017.30"
                        # Standard rounding should work.
                        
                        # Apply number format to ensure it displays 2 decimal places if needed
                        # cell.number_format = '0.00' # Optional: enforces display format

        output_filename = os.path.splitext(filename)[0] + "_Rounded.xlsx"
        print(f"Saving to: {output_filename}")
        wb.save(output_filename)
        print("Done.")
        return output_filename

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

if __name__ == "__main__":
    input_file = "Global Income Distribution Dataset.xlsx"
    # Full path is cleaner to ensure it finds it
    base_path = r"c:\Users\onkar\Desktop\infosys springboard virtual internship (1)"
    full_input_path = os.path.join(base_path, input_file)
    
    round_excel_values(full_input_path)
